# encoding: utf-8

import openpyxl
from PerformanceMeasure import PerformanceMeasure
import numpy as np
from network import *
from sklearn.utils import shuffle
from sklearn.preprocessing import MinMaxScaler
import torch
from loss import DAN
import torch.utils.data as data
import random

import time

import pandas as pd
class HuberLoss(nn.Module):
    def __init__(self, delta):
        super(HuberLoss, self).__init__()
        self.delta = delta

    def forward(self, y_true, y_pred):
        error = torch.abs(y_true - y_pred)
        quadratic = 0.5 * error ** 2
        linear = self.delta * (error - 0.5 * self.delta)
        loss = torch.where(error <= self.delta, quadratic, linear)
        return loss.mean()

def setup_seed(seed):
    torch.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    np.random.seed(seed)
    random.seed(seed)
    # 如果使用了 GPU，还可以设置 GPU 的随机数种子
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = True


class Classifier(nn.Module):
    def __init__(self, base_model, num_classes):
        super(Classifier, self).__init__()
        self.base_model = base_model
        self.classifier = nn.Linear(base_model.output_num(), num_classes)

    def forward(self, x):
        x = self.base_model(x)
        x = self.classifier(x)
        return x
def build_model(model_name, num_classes):
    base_model = network_dict[model_name]()
    model = Classifier(base_model, num_classes)
    return model


import torch.nn as nn


class Classifier(nn.Module):
    def __init__(self, base_model, num_classes):
        super(Classifier, self).__init__()
        self.base_model = base_model
        self.classifier = nn.Linear(base_model.output_num(), num_classes)

    def forward(self, x):
        x = self.base_model(x)
        x = self.classifier(x)
        return x


def expand_features_to_3_channels(features):
    # Reshape from [1090, 19] to [1090, 1, 19, 1]
    features = features.reshape(features.shape[0], 1, features.shape[1], 1)

    # Repeat the single channel three times to simulate RGB image
    features = np.repeat(features, 3, axis=1)

    return features


class SimpleDataset(data.Dataset):
    def __init__(self, features, labels=None):
        self.features = features
        self.labels = labels

    def __len__(self):
        return len(self.features)

    def __getitem__(self, idx):
        if self.labels is None:
            return self.features[idx]
        else:
            return self.features[idx], self.labels[idx]


def integrated_train(network_dict, model_name, source_features, source_labels, target_features, lr=0.001,
                     weight_decay=0.0001, num_epochs=10, batch_size=32):
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    source_features = expand_features_to_3_channels(source_features)
    target_features = expand_features_to_3_channels(target_features)
    source_labels = torch.tensor(source_labels, dtype=torch.float32).unsqueeze(1).to(device)
    source_features = torch.tensor(source_features, dtype=torch.float32)
    target_features = torch.tensor(target_features, dtype=torch.float32)

    # 使用 DataLoader
    source_dataset = SimpleDataset(source_features, source_labels)
    target_dataset = SimpleDataset(target_features)
    source_loader = data.DataLoader(source_dataset, batch_size=batch_size, shuffle=True)
    target_loader = data.DataLoader(target_dataset, batch_size=batch_size, shuffle=True)

    # Build Model
    base_model = network_dict[model_name]()
    model = Classifier(base_model, 1)
    model.to(device)

    optimizer = torch.optim.Adam(model.parameters(), lr=lr, weight_decay=weight_decay)
    class_criterion = HuberLoss(delta=1.0)
    transfer_criterion = DAN

    for epoch in range(num_epochs):
        model.train()
        for batch_source_features, batch_source_labels in source_loader:
            batch_target_features = next(iter(target_loader))
            optimizer.zero_grad()

            # Concatenate source and target features
            inputs = torch.cat((batch_source_features, batch_target_features), dim=0).to(device)
            outputs = model(inputs)

            classifier_loss = class_criterion(outputs[:len(batch_source_labels)], batch_source_labels.view(-1, 1))
            features = model.base_model(inputs)
            transfer_loss = transfer_criterion(features[:len(batch_source_labels)], features[len(batch_source_labels):])

            # Combine classification and transfer loss
            loss = classifier_loss + transfer_loss

            # Backward pass and optimization
            loss.backward()
            optimizer.step()

        # Print training progress
        if (epoch + 1) % 10 == 0:
            print(f"Epoch [{epoch + 1}/{num_epochs}], Loss: {loss.item():.4f}")

    return model



def train(source, target, seed):
    # Load Source Data
    cols = ['wmc', 'dit', 'noc', 'cbo', 'rfc', 'lcom', 'ca', 'ce', 'npm', 'lcom3', 'dam', 'moa', 'mfa', 'cam', 'ic',
            'cbm', 'amc', 'max_cc', 'avg_cc']
    source_file_path = f'../data/promise_csv/{source}.csv'
    source_data = pd.read_csv(source_file_path, usecols=cols)
    source_data = shuffle(source_data, random_state=seed)
    source_features = source_data.iloc[:, :].values
    label_data = pd.read_csv(source_file_path, usecols=['bug'])
    source_labels = label_data.iloc[:].values.ravel()

    # Load Target Data
    target_file_path = f'../data/promise_csv/{target}.csv'
    target_data = pd.read_csv(target_file_path, usecols=cols)
    target_features = target_data.iloc[:, :].values
    label_data = pd.read_csv(target_file_path, usecols=['bug'])
    loc_data = pd.read_csv(target_file_path, usecols=['loc'])
    loc_labels = loc_data.iloc[:].values.flatten()
    target_labels = label_data.iloc[:].values.flatten()



    source_features = np.log(source_features + 1e-6)
    target_features = np.log(target_features + 1e-6)


    # Feature Scaling
    scaler = MinMaxScaler().fit(source_features)
    source_features = scaler.transform(source_features)
    target_features = scaler.transform(target_features)

    # Train the model using integrated_train
    model = integrated_train(network_dict, "ResNet152", source_features, source_labels, target_features)

    # Predict using the model and calculate MSE
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    target_features = expand_features_to_3_channels(target_features)
    with torch.no_grad():
        model.eval()
        predictions = model(torch.tensor(target_features, dtype=torch.float32).to(device)).cpu().numpy()
    per = PerformanceMeasure(target_labels, predictions.flatten(), loc_labels)
    pofb = per.getPofb()

    # Return the MSE
    return pofb

if __name__ == "__main__":
    path = '../data/txt_png_path/'
    strings = ["ant-1.3", "camel-1.6", "ivy-2.0", "jedit-4.1", "log4j-1.2", "poi-2.0", "velocity-1.4", "xalan-2.4",
               "xerces-1.2"]
    new_arr = []
    test_arr = []
    # 构建所有可能的source和target组合
    for i in range(len(strings)):
        for j in range(i + 1, len(strings)):
            new_arr.append(strings[i] + "->" + strings[j])
            new_arr.append(strings[j] + "->" + strings[i])



    # 执行30次循环
    for round_cir in range(20):
        setup_seed(round_cir + 1)
        test_arr = []
        for str_combination in new_arr:
            source, target = str_combination.split("->")
            test_result = train(source, target, round_cir+1)
            print(f"{str_combination} test {test_result}")
            test_arr.append(float(test_result))

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for i in range(len(new_arr)):
            worksheet.cell(row=i + 1, column=1, value=new_arr[i])
            worksheet.cell(row=i + 1, column=2, value=test_arr[i])

        workbook.save('../output/average_originData_res152/' + str(round_cir + 1) + '_round.xlsx')  # 保存的文件名也修改为对

    # 计算平均值


