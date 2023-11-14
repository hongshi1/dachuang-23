# encoding: utf-8
import argparse
import os
from encoder import *
import torch.utils.data as data
from sklearn.metrics import mean_squared_error as mse
import cluster_AP
import numpy as np
import torch
import torch.nn as nn
import openpyxl
import cluster_spectral
# 定义和训练神经网络的类和函数。它包括各种层、激活函数、损失函数和配置神经网络结构的实用工具。
import torch.optim as optim
# 这个模块提供了用于在训练过程中更新神经网络模型权重的优化算法。
import network  # 引入自定义文件network.py
import loss
import pre_process as prep
# 这三个是自己定义的
import torch.utils.data as util_data  # To use 'DataLoader()'
import lr_schedule
from data_list import ImageList
from torch.autograd import Variable
from PerformanceMeasure import Origin_PerformanceMeasure as PerformanceMeasure
# 貌似已经被弃用，主要是为了允许在安详传播的过程中进行自动微分来计算梯度
import math

optim_dict = {"ADAM": optim.Adam, "SGD": optim.SGD}
from sklearn.manifold import TSNE
import matplotlib.pyplot as plt
import random
import time
from PIL import Image


# 主训练函数

# def standardize_batch(features):
#     """
#     Standardize the features by removing the mean and scaling to unit variance
#     """
#     # mean = features.mean(dim=0, keepdim=True)
#     # std = features.std(dim=0, keepdim=True) + 1e-6  # 防止除以0
#     # features_standardized = (features - mean) / std
#     # return features_standardized
#     min_val = features.min(dim=0, keepdim=True)[0]
#     max_val = features.max(dim=0, keepdim=True)[0]
#
#     # Avoid division by zero by adding a small constant (1e-6)
#     range_val = max_val - min_val + 1e-6
#
#     features_normalized = (features - min_val) / range_val
#     return features_normalized


def process_data(data, device):
    """
    Process a single batch of data.
    """
    labels = data[1].to(device)
    imgName = data[2]
    astVec = data[3].to(device)  # AST vector.
    imgVec = data[4].squeeze(1).to(device)  # Image vector, reshaped as needed.

    # Get a subset of labels (excluding the first column)
    labels_subset = labels[:, 1:]

    # Concatenate the AST vector, Image vector, and the subset of labels.
    combinedVec = torch.cat((astVec, imgVec, labels_subset), dim=1).to(torch.float32)
    return combinedVec, labels


def compute_features_and_loss(iter_source, iter_target, base_network, regressor_layer, class_criterion,
                              transfer_criterion, loss_config, device, net_config, bottleneck_layer=None):
    """
    Compute features from source and target, then compute the loss.
    """
    # Process source data
    data_source = next(iter_source)
    combinedVec_s, labels_source = process_data(data_source, device)
    # combinedVec_s = standardize_batch(combinedVec_s)

    features_source = base_network(combinedVec_s)

    # Process target data
    data_target = next(iter_target)
    combinedVec_t, labels_target = process_data(data_target, device)
    # combinedVec_t = standardize_batch(combinedVec_t)
    features_target = base_network(combinedVec_t)

    # Combine the features
    features_combined = torch.cat((features_source, features_target), dim=0)

    # Apply bottleneck layer if configured
    if net_config.get("use_bottleneck"):
        features_combined = bottleneck_layer(features_combined)

    # Compute the regressor output
    # outputs = regressor_layer(features_combined)

    output_s = regressor_layer(features_source)
    p = PerformanceMeasure(labels_target[:, 0].cpu(), output_s.detach().cpu(), labels_target[:, 1].cpu(),
                           labels_target[:, 20].cpu())
    popt = p.PercentPOPT().to(device)
    bug_s = labels_source[:, 0].float().view(-1, 1)

    # Compute the regressor loss using the source data
    regressor_loss = class_criterion( bug_s,output_s)

    # Compute the transfer loss
    transfer_loss = compute_transfer_loss(features_combined, transfer_criterion, loss_config)

    return regressor_loss, transfer_loss


def compute_transfer_loss(features_combined, transfer_criterion, loss_config):
    """
    Compute the transfer loss based on the loss configuration.
    """
    half_size = len(features_combined) // 2
    source_features = features_combined[:half_size]
    target_features = features_combined[half_size:]

    if loss_config["name"] == "DAN":
        transfer_loss = transfer_criterion(source_features, target_features, **loss_config["params"])
    elif loss_config["name"] == "RTN":
        # RTN is still under development
        transfer_loss = 0
    elif loss_config["name"] == "JAN":
        # JAN requires softmax outputs as well
        # softmax_out = softmax_layer(outputs)
        # source_softmax = softmax_out[:half_size]
        # target_softmax = softmax_out[half_size:]
        # transfer_loss = transfer_criterion(
        #     [source_features, source_softmax],
        #     [target_features, target_softmax],
        #     **loss_config["params"]
        # )
        pass
    else:
        raise ValueError("Unknown transfer loss name: {}".format(loss_config["name"]))

    return transfer_loss


class HuberLoss(nn.Module):
    def __init__(self, delta):
        super(HuberLoss, self).__init__()
        self.delta = delta

    def forward(self, y_true, y_pred):
        # 对 y_pred 中大于 20 的部分应用 log100 映射
        y_pred_mapped = torch.where(y_pred > 20, 20 + torch.log(y_pred - 20 + 1) / torch.log(torch.tensor(100.0)), y_pred)

        error = torch.abs(y_true - y_pred_mapped)
        quadratic = 0.5 * error ** 2
        linear = self.delta * (error - 0.5 * self.delta)
        loss = torch.where(error <= self.delta, quadratic, linear)
        return loss.mean()


def get_fea_lab(what, loader, model, gpu):
    start_test = True
    iter_what = iter(loader[what])
    for i in range(len(loader[what])):
        data = iter_what.next()
        inputs = data[0]
        labels = data[1]
        if gpu:
            inputs = Variable(inputs.cuda())
            labels = Variable(labels.cuda())
        else:
            inputs = Variable(inputs)
            labels = Variable(labels)

        outputs = model(inputs)  # get features

        if start_test:
            all_output = outputs.data.float()
            all_label = labels.data.float()
            start_test = False
        else:
            all_output = torch.cat((all_output, outputs.data.float()), 0)
            all_label = torch.cat((all_label, labels.data.float()), 0)

    # all_label_list = all_label.view(-1.1).cpu().numpy()
    # all_output_list = all_output.view(-1,1).cpu().numpy()
    all_label_list = all_label.cpu().numpy()
    all_output_list = all_output.cpu().numpy()
    return all_output_list, all_label_list


def setup_seed(seed):
    torch.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    np.random.seed(seed)
    random.seed(seed)
    # 如果使用了 GPU，还可以设置 GPU 的随机数种子
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = True


def tsne(df_fea, df_lab):
    # print(df_fea, df_lab)
    tsne = TSNE(2, 38, 20, 600)
    result = tsne.fit_transform(df_fea)
    colors = ['c', 'r']
    idx_1 = [i1 for i1 in range(len(df_lab)) if df_lab[i1] == 0]
    fig1 = plt.scatter(result[idx_1, 0], result[idx_1, 1], 20, color=colors[0], label='Clean')

    idx_2 = [i2 for i2 in range(len(df_lab)) if df_lab[i2] == 1]
    fig2 = plt.scatter(result[idx_2, 0], result[idx_2, 1], 20, color=colors[1], label='Defective')


def get_tsne_img(what, loader, model, gpu=True):
    fea, lab = get_fea_lab(what, loader, model, gpu)
    tsne(fea, lab)
    plt.legend()
    plt.xticks([])
    plt.yticks([])

    if what == 'train':
        plt.savefig(
            '../temp/tsne/' + args.source + '-' + args.target + '.pdf')
    else:
        plt.savefig(
            '../temp/tsne/' + args.target + '-' + args.source + '.pdf')
    plt.close()


def image_classification_predict(loader, model, test_10crop=False, gpu=True):
    start_test = True
    device = torch.device("cuda:0" if torch.cuda.is_available() and gpu else "cpu")

    model = model.to(device)

    iter_test = iter(loader["test"])
    for _ in range(len(loader["test"])):
        data = next(iter_test)
        inputs = data[0].to(device)  # This is the image data.
        labels = data[1].to(device)
        imgName = data[2]
        astVec = data[3].to(device)  # AST vector.
        imgVec = data[4].squeeze(1).to(device)  # Image vector, reshaped as needed.

        labels = Variable(labels)

        labels = Variable(labels)
        labels_subset = labels[:, 1:]

        # Concatenate the AST vector, Image vector, and the subset of labels.
        combinedVec = torch.cat((astVec, imgVec, labels_subset), dim=1).to(torch.float32)
        # combinedVec = standardize_batch(combinedVec)

        # 待定
        outputs = model(combinedVec)

        if start_test:
            all_output = outputs.data.float()
            all_label = labels.data.float()
            start_test = False
        else:
            all_output = torch.cat((all_output, outputs.data.float()), 0)
            all_label = torch.cat((all_label, labels.data.float()), 0)

    predict = all_output.flatten()
    return all_label, predict


# I'll continue the modifications for the second function here:

def image_classification_test(loader, model, test_10crop=False, gpu=True):
    start_test = True
    device = torch.device("cuda:0" if torch.cuda.is_available() and gpu else "cpu")

    model = model.to(device)

    iter_test = iter(loader["test"])
    for _ in range(len(loader["test"])):
        data = next(iter_test)
        inputs = data[0].to(device)  # 指的是图片
        labels = data[1].to(device)
        imgName = data[2]
        astVec = data[3].to(device)
        imgVec = data[4].squeeze(1).to(device)
        # 各种各样的label输入 第一行是代码bug,第二行是loc 等等等

        labels = Variable(labels)
        labels_subset = labels[:, 1:]

        # Concatenate the AST vector, Image vector, and the subset of labels.
        combinedVec = torch.cat((astVec, imgVec, labels_subset), dim=1).to(torch.float32)
        # combinedVec = standardize_batch(combinedVec)

        # 待定
        outputs = model(combinedVec)

        if start_test:
            all_output = outputs.data.float()
            all_label = labels.data.float()
            start_test = False
        else:
            all_output = torch.cat((all_output, outputs.data.float()), 0)
            all_label = torch.cat((all_label, labels.data.float()), 0)

    predict_list = all_output.round().cpu().numpy().flatten()
    all_label_list = all_label.cpu().numpy()
    popt = -1.0
    loc = all_label_list[:, 1]
    cc = all_label_list[:, 20]

    if (all_label_list.shape[1] > 1):
        p = PerformanceMeasure(all_label_list[:, 0], predict_list, loc, cc)
        popt = p.PercentPOPT()

    return popt


def transfer_classification(config):
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    # 定义一个字典类型变量
    prep_dict = {}
    # Add kry-value pairs for 'prep_dict'
    # 数据预处理
    for prep_config in config["prep"]:
        prep_dict[prep_config["name"]] = {}
        if prep_config["type"] == "image":
            prep_dict[prep_config["name"]]["test_10crop"] = prep_config["test_10crop"]
            # Call image_train() in pre_process.py
            prep_dict[prep_config["name"]]["train"] = prep.image_train(resize_size=prep_config["resize_size"],
                                                                       crop_size=prep_config["crop_size"])
            # call image_test()

            if prep_config["test_10crop"]:
                prep_dict[prep_config["name"]]["test"] = prep.image_test_10crop(resize_size=prep_config["resize_size"],
                                                                                crop_size=prep_config["crop_size"])
            else:
                prep_dict[prep_config["name"]]["test"] = prep.image_test(resize_size=prep_config["resize_size"],
                                                                         crop_size=prep_config["crop_size"])

    # class_criterion = nn.CrossEntropyLoss()         ##交叉熵损失函数
    # class_criterion = nn.CrossEntropyLoss(weight=torch.tensor([1, 1, 2]))

    class_criterion = HuberLoss(1.0)
    # class_criterion = nn.MSELoss()
    loss_config = config["loss"]
    # 如果在配置文件的loss部分中，name属性指定为DAN，那么表示使用DAN（Domain Adversarial Neural Networks）方法来进行域适应（domain adaptation）学习。
    # DAN是一种常用的域适应方法，它通过对抗训练的方式来使得特征提取器对源域和目标域的特征表示具有相同的分布，从而提高模型的泛化性能。在具体实现中，DAN使用一个域分类器来判
    # 别输入的特征属于源域还是目标域，同时通过最小化这个分类器的损失来训练特征提取器，使得提取到的特征能够欺骗域分类器，使得域分类器无法判断输入的特征是来自源域还是目标域。
    transfer_criterion = loss.loss_dict[loss_config["name"]]
    if "params" not in loss_config:
        loss_config["params"] = {}

    ## prepare data
    dsets = {}
    dset_loaders = {}
    for data_config in config["data"]:
        dsets[data_config["name"]] = {}  # 创建字典元素
        dset_loaders[data_config["name"]] = {}  # 创建一个字典元素
        if data_config["type"] == "image":
            # Creat a 'ImageList' object，读取txt文件中的数据赋值给字典元素'train'
            dsets[data_config["name"]]["train"] = ImageList(open(data_config["list_path"]["train"]).readlines(),
                                                            transform=prep_dict[data_config["name"]]["train"])
            # 创建一个DataLoader对象，为字典'dset_loaders'中嵌套字典元素'train'赋值
            dset_loaders[data_config["name"]]["train"] = util_data.DataLoader(dsets[data_config["name"]]["train"],
                                                                              batch_size=data_config["batch_size"][
                                                                                  "train"], shuffle=True, num_workers=4)
            if "test" in data_config["list_path"]:
                if prep_dict[data_config["name"]]["test_10crop"]:
                    for i in range(10):  # 10?
                        dsets[data_config["name"]]["test" + str(i)] = ImageList(
                            open(data_config["list_path"]["test"]).readlines(),
                            transform=prep_dict[data_config["name"]]["test"]["val" + str(i)]
                        )
                        dset_loaders[data_config["name"]]["test" + str(i)] = util_data.DataLoader(
                            dsets[data_config["name"]]["test" + str(i)], batch_size=data_config["batch_size"]["test"],
                            shuffle=False, num_workers=4)
                else:
                    dsets[data_config["name"]]["test"] = ImageList(open(data_config["list_path"]["test"]).readlines(),
                                                                   transform=prep_dict[data_config["name"]]["test"])

                    dset_loaders[data_config["name"]]["test"] = util_data.DataLoader(dsets[data_config["name"]]["test"],
                                                                                     batch_size=
                                                                                     data_config["batch_size"]["test"],
                                                                                     shuffle=False, num_workers=4)
            else:
                if prep_dict[data_config["name"]]["test_10crop"]:
                    for i in range(10):
                        dsets[data_config["name"]]["test" + str(i)] = ImageList(
                            open(data_config["list_path"]["train"]).readlines(),
                            transform=prep_dict[data_config["name"]]["test"]["val" + str(i)])
                        dset_loaders[data_config["name"]]["test" + str(i)] = util_data.DataLoader(
                            dsets[data_config["name"]]["test" + str(i)], batch_size=data_config["batch_size"]["test"],
                            shuffle=False, num_workers=4)
                else:
                    dsets[data_config["name"]]["test"] = ImageList(open(data_config["list_path"]["train"]).readlines(),
                                                                   transform=prep_dict[data_config["name"]]["test"])
                    dset_loaders[data_config["name"]]["test"] = util_data.DataLoader(dsets[data_config["name"]]["test"],
                                                                                     batch_size=
                                                                                     data_config["batch_size"]["test"],
                                                                                     shuffle=False, num_workers=4)

    class_num = 1  # ??

    ## set base network
    net_config = config["network"]
    bottleneck_layer = ''
    base_network = network.network_dict[net_config["name"]]()  # 'network_dict'是一个字典，包含各种类型的AlexNet
    if net_config["use_bottleneck"]:
        bottleneck_layer = nn.Linear(base_network.output_num(), net_config["bottleneck_dim"])  # 创建瓶颈层
        regressor_layer = nn.Linear(bottleneck_layer.out_features, 1, bias=True)  # 创建回归层
    else:
        regressor_layer = nn.Linear(base_network.output_num(), 1, bias=True)  # 创建回归层
    for param in base_network.parameters():
        param.requires_grad = True

    ## initialization
    if net_config["use_bottleneck"]:
        bottleneck_layer.weight.data.normal_(0, 0.005)
        bottleneck_layer.bias.data.fill_(0.1)
        bottleneck_layer = nn.Sequential(bottleneck_layer, nn.ReLU(), nn.Dropout(0.6))

    use_gpu = torch.cuda.is_available()
    print(use_gpu)
    if use_gpu:  # 如果GPU可用则不使用CPU进行训练，默认使用CPU
        if net_config["use_bottleneck"]:
            bottleneck_layer = bottleneck_layer.cuda()
        regressor_layer = regressor_layer.cuda()
        base_network = base_network.cuda()

    ## collect parameters
    if net_config["use_bottleneck"]:
        parameter_list = [{"params": base_network.parameters(), "lr": 0.1},
                          {"params": bottleneck_layer.parameters(), "lr": 0.1},
                          {"params": regressor_layer.parameters(), "lr": 0.1}]
    else:
        parameter_list = [{"params": base_network.parameters(), "lr": 0.01},
                          {"params": regressor_layer.parameters(), "lr": 0.01}]

    ## add additional network for some methodsf
    if loss_config["name"] == "JAN":
        softmax_layer = nn.Softmax()
        if use_gpu:
            softmax_layer = softmax_layer.cuda()

    nn.init.normal_(regressor_layer.weight, mean=0.0, std=0.01)
    nn.init.zeros_(regressor_layer.bias)

    ## set optimizer
    optimizer_config = config["optimizer"]
    optimizer = optim_dict[optimizer_config["type"]](parameter_list, **(optimizer_config["optim_params"]))
    param_lr = []
    for param_group in optimizer.param_groups:
        param_lr.append(param_group["lr"])

    ## train
    len_train_source = len(dset_loaders["source"]["train"]) - 1
    len_train_target = len(dset_loaders["target"]["train"]) - 1
    F_best = 0  # F-measure的取值范围是[0,1]，值越小表示模型性能越差，所以其最优值初始化为0

    best_model = ''
    predict_best = ''
    for i in range(config["num_iterations"]):  # 网格法确定最佳参数组合
        if F_best >= 1:
            break
        else:
            if i % config["test_interval"] == 0:  # "test_interval"?
                base_network.train(False)
                regressor_layer.train(False)  # False -- ?
                if net_config["use_bottleneck"]:
                    bottleneck_layer.train(False)
                    F = image_classification_test(dset_loaders["source"],  # not 'target' when training
                                                  nn.Sequential(base_network, bottleneck_layer, regressor_layer),
                                                  test_10crop=prep_dict["source"]["test_10crop"], gpu=use_gpu)
                else:
                    F = image_classification_test(dset_loaders["source"],  # not 'target' when training
                                                  nn.Sequential(base_network, regressor_layer),
                                                  # nn.Sequential一个用于存放神经网络模块的序列容器，可以用来自定义模型，运行顺序按照输入顺序进行
                                                  test_10crop=prep_dict["source"]["test_10crop"], gpu=use_gpu)

                print(args.source + '->' + args.target)
                print("F")
                print(F)
                if F_best < F:
                    F_best = F
                    base_network.train(False)
                    regressor_layer.train(False)
                    if net_config["use_bottleneck"]:
                        bottleneck_layer.train(False)
                        best_model = nn.Sequential(base_network, bottleneck_layer, regressor_layer)
                        all_label, predict_best = image_classification_predict(dset_loaders["target"], best_model,
                                                                               test_10crop=False, gpu=use_gpu)
                    else:
                        best_model = nn.Sequential(base_network, regressor_layer)
                        all_label, predict_best = image_classification_predict(dset_loaders["target"], best_model,
                                                                               test_10crop=False, gpu=use_gpu)

            loss_test = nn.BCELoss()
            ## train one iter
            if net_config["use_bottleneck"]:
                bottleneck_layer.train(True)
            regressor_layer.train(True)  # 将模型设置为训练模式
            # optimizer_config = config["optimizer"]
            # optimizer = optim_dict[optimizer_config["type"]](parameter_list, **(optimizer_config["optim_params"]))
            # 调整优化器的学习率，学习率调度程序有StepLR，MultiStepLR，ExponentialLR等，param_lr是一个包含每个参数组初始学习率的列表，optimizer是优化器，i是当前迭代次数，schedule_param包含调度程序的参数
            optimizer.zero_grad()  # 用于将梯度缓存清零
            if i % len_train_source == 0:
                iter_source = iter(dset_loaders["source"]["train"])  # 更新源域数据集迭代器
            if i % len_train_target == 0:
                iter_target = iter(dset_loaders["target"]["train"])  # 更新目标域数据集迭代器

            base_network = base_network.to(device)
            # bottleneck_layer = bottleneck_layer.to(device)
            regressor_layer = regressor_layer.to(device)

            regressor_loss, transfer_loss = compute_features_and_loss(
                iter_source,
                iter_target,
                base_network,
                regressor_layer,
                class_criterion,
                transfer_criterion,
                loss_config,
                device,
                net_config,
                bottleneck_layer,
            )

            rate = config["distances"][config["clusters"][args.source]][config["clusters"][args.target]]
            # total_loss = 1 * transfer_loss + classifier_loss
            total_loss = regressor_loss
            print("regressor_loss:", total_loss.item())
            print("transfer_loss:", transfer_loss.item())
            #
            end_train = time.perf_counter()

            # print('loss: %.4f' % total_loss)
            total_loss.backward()
            optimizer.step()

    print(args.source + '->' + args.target)
    print('训练结果：')
    print(F_best)
    popt = 0.0

    all_label_list = all_label.cpu().numpy()
    predict_list = predict_best.view(-1, 1).round().cpu().numpy().flatten()
    loc = all_label_list[:, 1]
    cc = all_label_list[:, 20]

    if (all_label_list.shape[1] > 1):
        p = PerformanceMeasure(all_label_list[:, 0], predict_list, loc, cc)
        popt = p.PercentPOPT()
    print(popt)
    return popt


if __name__ == "__main__":
    # random.seed(time.time())
    # setup_seed(random.randint(1, 100))
    setup_seed(20)
    # path = '..\data\img\grb_img\ivy-2.0\\buggy\ivy-2.0_src_java_org_apache_ivy_core_IvyPatternHelper.png'
    # # path = 'F:\Document\GitHub\DTLDP_master\data\img\grb_img\ivy-2.0\clean\ivy-2.0_src_java_org_apache_ivy_ant_AddPathTask.png'
    # with open(path, 'rb') as f:  # 以二进制格式打开一个文件用于只读
    #     with Image.open(f) as img:
    #         a = img.convert('RGB')

    path = '../data/txt_png_path/'
    # path = '../data/txt/'

    # Case1: 使用命令行
    # # 创建一个解析对象，对命令行参数进行解析 （这种方式不利于调试，可以对args进行赋值）
    # parser = argparse.ArgumentParser(description='Transfer Learning')
    # # 添加参数
    # parser.add_argument('--gpu_id', type=str, nargs='?', default='0', help="device id to run")
    # parser.add_argument('--source', type=str, nargs='?', default='ant-1.6', help="source data")
    # parser.add_argument('--target', type=str, nargs='?', default='ant-1.7', help="target data")
    # parser.add_argument('--loss_name', type=str, nargs='?', default='DAN', help="loss name")
    # parser.add_argument('--tradeoff', type=float, nargs='?', default=1.0, help="tradeoff")
    # parser.add_argument('--using_bottleneck', type=int, nargs='?', default=1, help="whether to use bottleneck")
    # parser.add_argument('--task', type=str, nargs='?', default='WPDP', help="WPDP or CPDP")
    # # 解析
    # args = parser.parse_args()

    # Case2: 不使用命令行
    strings = ["ant-1.3", "camel-1.6", "ivy-2.0", "jedit-4.1", "log4j-1.2", "poi-2.0", "velocity-1.4", "xalan-2.4",
               "xerces-1.2"]
    # strings = ["ant-1.3", "velocity-1.4"]
    new_arr = []
    test_arr = []

    for i in range(len(strings)):
        for j in range(i + 1, i + 2):
            m = (i + 1) % len(strings)
            n = (i + 2) % len(strings)
            new_arr.append(strings[i] + "->" + strings[m])
            new_arr.append(strings[i] + "->" + strings[n])

    parser = argparse.ArgumentParser(description='Transfer Learning')
    args = parser.parse_args()
    args.gpu_id = '0'
    args.source = 'xalan-2.4'
    args.target = 'xalan-2.4'
    args.loss_name = 'DAN'
    args.tradeoff = 1.0
    args.using_bottleneck = 0
    args.task = 'CPDP'  # 'WPDP' or 'CPDP'
    # cpdp 表示跨项目缺陷预测

    # kmeans++聚类
    # clusters, distances = cluster.project_cluster(3)
    # 谱聚类
    # clusters, distances = cluster_spectral.project_cluster(3)
    clusters, distances = cluster_AP.project_cluster()
    for round_cir in range(20):
        new_arr = []
        test_arr = []

        for i in range(len(strings)):
            for j in range(i + 1, i + 2):
                m = (i + 1) % len(strings)
                n = (i + 2) % len(strings)
                new_arr.append(strings[i] + "->" + strings[m])
                new_arr.append(strings[i] + "->" + strings[n])

        for i in range(len(new_arr)):
            setup_seed(round_cir + 1)
            args.source = new_arr[i].split("->")[0]
            args.target = new_arr[i].split("->")[1]
            mytarget_path = "../data/txt/" + args.target + ".txt"
            classnum = 1
            print(args.source + " " + args.target + " ", end='')
            print(classnum)

            os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu_id

            # 定义一个字典类型变量
            config = {}
            # 添加键值对
            config["num_iterations"] = 20
            config["test_interval"] = 1  # ?
            # test_10crop 是一个布尔类型的参数，用于表示在测试集上是否进行 10-crop 测试。10-crop 测试是指在测试时将一张图片切成 10 个部分并对每个部分进行预测，然后将这 10 个预测结果进行平均或投票得到最终的预测结果。这种方法可以提高模型的准确性，特别是在处理图像数据时。
            config["prep"] = [
                {"name": "source", "type": "image", "test_10crop": False, "resize_size": 256, "crop_size": 224},
                {"name": "target", "type": "image", "test_10crop": False, "resize_size": 256, "crop_size": 224}]
            config["loss"] = {"name": args.loss_name, "trade_off": args.tradeoff}
            #
            config["data"] = [{"name": "source", "type": "image", "list_path": {"train": path + args.source + ".txt"},
                               "batch_size": {"train": 32, "test": 32}},
                              {"name": "target", "type": "image", "list_path": {"train": path + args.target + ".txt"},
                               "batch_size": {"train": 32, "test": 32}}]
            config["network"] = {"name": "dpnn", "use_bottleneck": args.using_bottleneck,
                                 "bottleneck_dim": 256}
            # config["optimizer"] = {"type": "SGD",
            #                        "optim_params": {"lr": 0.005, "momentum": 0.9, "weight_decay": 0.05,
            #                                         "nesterov": True},
            #                        "lr_type": "inv", "lr_param": {"init_lr": 0.0001, "gamma": 0.0003, "power": 0.75}}

            config["clusters"] = clusters
            config["distances"] = distances
            # config["rate"] = [5, 10, 100]
            config["optimizer"] = {
                "type": "ADAM",
                "optim_params": {"lr": 0.0001, "betas": (0.9, 0.999), "eps": 1e-08, "weight_decay": 0.0005,
                                 "amsgrad": False},
                "lr_type": "inv", "lr_param": {"init_lr": 0.0001, "gamma": 0.06, "power": 0.6}
            }

            # 对代码的修改和理解  都吧注释写满  方便组员学习
            # num_iterations表示训练的迭代次数；
            # test_interval表示每多少个迭代进行一次测试；
            # prep表示数据预处理的配置，包括source和target两个来源的数据，需要进行的操作包括图片的缩放和裁剪；
            # loss表示损失函数的配置，包括使用的损失函数的名称和对各项损失的权重；
            # data表示训练和测试数据的配置，包括source和target两个来源的数据，需要读取的文件路径和每个batch的大小；
            # network表示神经网络的配置，包括使用的网络名称、是否使用bottleneck特征、bottleneck的维度等；
            # optimizer表示优化器的配置，包括使用的优化算法、学习率、动量、权重衰减等参数。
            test_result = transfer_classification(config)
            print(new_arr[i], end=' ')
            print(" popt_final", end=' ')
            print(test_result)
            test_arr.append(test_result)

        workbook = openpyxl.Workbook()
        # 选择默认的工作表
        worksheet = workbook.active

        for i in range(len(new_arr)):
            worksheet.cell(row=i + 1, column=1, value=new_arr[i])
            worksheet.cell(row=i + 1, column=2, value=test_arr[i])
        # 保存文件
        workbook.save('../output/newloss_round/' + str(round_cir + 1) + '_adam_round.xlsx')  # 运行失败 需要改一个别的文件名