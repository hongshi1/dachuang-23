# encoding: utf-8

import openpyxl
import random
import time
import openpyxl
from PerformanceMeasure import Origin_PerformanceMeasure as PerformanceMeasure
import random
import time
import pandas as pd
from sklearn.metrics import mean_squared_error
from sklearn.tree import DecisionTreeRegressor
import numpy as np
from sklearn.preprocessing import MinMaxScaler
import torch
import torch.nn as nn
import torch.optim as optim


# Define your deep regressor model
class DeepRegressor(nn.Module):
    def __init__(self, input_dim=20):
        super(DeepRegressor, self).__init__()
        self.fc1 = nn.Linear(input_dim, 64)
        self.fc2 = nn.Linear(64, 32)
        self.fc3 = nn.Linear(32, 1)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = self.fc3(x)
        return x
class LinearRegressor(nn.Module):
    def __init__(self, input_dim=20):
        super(LinearRegressor, self).__init__()
        self.fc = nn.Linear(input_dim, 1)

    def forward(self, x):
        return self.fc(x)

class MLPRegressor(nn.Module):
    def __init__(self, input_dim, hidden_dim, num_layers):
        super(MLPRegressor, self).__init__()
        layers = []
        layers.append(nn.Linear(input_dim, hidden_dim))
        layers.append(nn.ReLU())
        for _ in range(num_layers - 1):
            layers.append(nn.Linear(hidden_dim, hidden_dim))
            layers.append(nn.ReLU())
        layers.append(nn.Linear(hidden_dim, 1))
        self.model = nn.Sequential(*layers)

    def forward(self, x):
        return self.model(x)


# Training function using the deep regressor model
def train(source, target):
    # Load Source Data
    cols = ['wmc', 'dit', 'noc', 'cbo', 'rfc', 'lcom', 'ca', 'ce', 'npm', 'lcom3', 'dam', 'moa', 'mfa', 'cam', 'ic',
            'cbm', 'amc', 'max_cc', 'avg_cc','loc']
    source_file_path = f'../data/promise_csv/{source}.csv'
    source_data = pd.read_csv(source_file_path, usecols=cols)
    source_features = source_data.iloc[:, :].values
    label_data = pd.read_csv(source_file_path, usecols=['bug'])
    source_labels = label_data.iloc[:, :].values
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    # Load Target Data
    target_file_path = f'../data/promise_csv/{target}.csv'
    target_data = pd.read_csv(target_file_path, usecols=cols)
    target_features = target_data.iloc[:, :].values
    label_data = pd.read_csv(target_file_path, usecols=['bug'])
    loc_data = pd.read_csv(target_file_path, usecols=['loc'])
    loc_labels = loc_data.iloc[:, :].values.flatten()
    target_labels = label_data.iloc[:, :].values.flatten()

    # Log Transformation
    source_features = np.log(source_features + 1e-6)
    target_features = np.log(target_features + 1e-6)

    # Feature Scaling (Normalization)
    scaler = MinMaxScaler().fit(source_features)
    source_features = scaler.transform(source_features)
    target_features = scaler.transform(target_features)

    # Create PyTorch tensors
    source_features = torch.Tensor(source_features)
    source_labels = torch.Tensor(source_labels)
    loc_data = pd.read_csv(target_file_path, usecols=['loc'])
    cc_data = pd.read_csv(target_file_path, usecols=['avg_cc'])  # Add this line to load "cc" feature
    cc_labels = cc_data.iloc[:].values.flatten()
    target_features = torch.Tensor(target_features)

    # Define your deep regressor model
    model = DeepRegressor(input_dim=source_features.shape[1])
    # model = MLPRegressor(input_dim=20, hidden_dim=64, num_layers=20).to(device)
    # Define loss function and optimizer
    criterion = nn.MSELoss()
    optimizer = optim.Adam(model.parameters(), lr=0.001)

    # Training loop
    for epoch in range(100):
        optimizer.zero_grad()
        outputs = model(source_features.to(device))
        loss = criterion(outputs.to(device), source_labels.to(device))
        # print("第", epoch, "轮：", "loss=", loss.item())
        loss.backward()
        optimizer.step()

    # Switch to evaluation mode
    model.eval()

    # Predict using the model
    with torch.no_grad():
        predictions = model(target_features.to(device))

    # Convert predictions to numpy array for the PerformanceMeasure class
    predictions = torch.round(predictions.cpu().numpy().flatten())

    # Calculate and return the POPT
    per = PerformanceMeasure(target_labels, predictions, loc_labels,cc_labels)
    popt = per.PercentPOPT()

    return popt


if __name__ == "__main__":
    path = '../data/txt_png_path/'
    strings = ["ant-1.3", "camel-1.6", "ivy-2.0", "jedit-4.1", "log4j-1.2", "poi-2.0", "velocity-1.4", "xalan-2.4",
               "xerces-1.2"]
    new_arr = []

    # Build all possible source and target combinations
    for i in range(len(strings)):
        for j in range(i + 1, len(strings)):
            new_arr.append(strings[i] + "->" + strings[j])
            new_arr.append(strings[j] + "->" + strings[i])

    all_test_results = []  # Store results for each iteration

    # Run 30 iterations
    for _ in range(30):
        random.seed(time.time())
        seed = random.randint(1, 100)
        test_arr = []

        for str_combination in new_arr:
            source, target = str_combination.split("->")
            test_result = train(source, target)
            print(f"{str_combination} test {test_result}")
            test_arr.append(float(test_result))

        all_test_results.append(test_arr)  # Append test results for the current iteration

    # Calculate the average results
    avg_test_results = [sum(x) / 30 for x in zip(*all_test_results)]

    # Create an Excel workbook and write the results
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for i, (combination, avg_result) in enumerate(zip(new_arr, avg_test_results)):
        worksheet.cell(row=i + 1, column=1, value=combination)
        worksheet.cell(row=i + 1, column=2, value=avg_result)

    # Save the Excel file
    workbook.save('../output/average_output_MLP_popt_30_log_perpopt_newData_loc_2.xlsx')
