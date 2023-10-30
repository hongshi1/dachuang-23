# encoding: utf-8

import openpyxl
from PerformanceMeasure  import Origin_PerformanceMeasure as PerformanceMeasure
import torch.optim as optim
import random
import time
import pandas as pd
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split
from sklearn.tree import DecisionTreeRegressor
import numpy as np
from sklearn.preprocessing import MinMaxScaler
#决策树的方法结果

def train(source, target):
    # Load Source Data
    cols = ['wmc','dit','noc','cbo','rfc','lcom','ca','ce','npm','lcom3','dam','moa','mfa','cam','ic','cbm','amc','max_cc','avg_cc']
    source_file_path = f'../data/promise_csv/{source}.csv'
    source_data = pd.read_csv(source_file_path, usecols=cols)  # Columns D to W are 3 to 22
    source_features = source_data.iloc[:,:].values  # All columns except the last one
    label_data = pd.read_csv(source_file_path, usecols=['bug'])
    source_labels = label_data.iloc[:].values  # The last column



    # Load Target Data
    target_file_path = f'../data/promise_csv/{target}.csv'
    target_data = pd.read_csv(target_file_path, usecols=cols)  # Columns D to W are 3 to 22
    target_features = target_data.iloc[:, :].values  # All columns except the last one
    label_data = pd.read_csv(target_file_path, usecols=['bug'])
    loc_data = pd.read_csv(target_file_path, usecols=['loc'])
    loc_labels = loc_data.iloc[:].values.flatten()
    target_labels = label_data.iloc[:].values.flatten()  # The last column

    # 1. Log Transformation
    source_features = np.log(source_features + 1e-6)
    target_features = np.log(target_features + 1e-6)

    # 2. Feature Scaling (Normalization)
    scaler = MinMaxScaler().fit(source_features)  # Only fit on source_features
    source_features = scaler.transform(source_features)
    target_features = scaler.transform(target_features)
    cc_data = pd.read_csv(target_file_path, usecols=['avg_cc'])  # Add this line to load "cc" feature
    cc_labels = cc_data.iloc[:].values.flatten()

    # Train a Decision Tree Model using Source Data
    model = DecisionTreeRegressor()
    model.fit(source_features, source_labels)

    # Predict using the model and calculate MSE
    predictions = model.predict(target_features)
    per = PerformanceMeasure(target_labels, predictions,loc_labels,cc_labels)

    pofb = per.PercentPOPT()


    # Return the MSE
    return pofb


if __name__ == "__main__":
    path = '../data/txt_png_path/'
    strings = ["ant-1.3", "camel-1.6", "ivy-2.0", "jedit-4.1", "log4j-1.2", "poi-2.0", "velocity-1.4", "xalan-2.4",
               "xerces-1.2"]
    new_arr = []

    # 构建所有可能的source和target组合
    for i in range(len(strings)):
        for j in range(i + 1, len(strings)):
            new_arr.append(strings[i] + "->" + strings[j])
            new_arr.append(strings[j] + "->" + strings[i])

    all_test_results = []  # 存储每次循环的所有测试结果

    # 执行30次循环
    for _ in range(30):
        random.seed(time.time())
        seed = random.randint(1, 100)
        test_arr = []
        for str_combination in new_arr:
            source, target = str_combination.split("->")
            test_result = train(source, target)
            print(f"{str_combination} test {test_result}")
            test_arr.append(float(test_result))
        all_test_results.append(test_arr)  # 添加当前循环的测试结果到总列表

    # 计算平均值
    avg_test_results = [sum(x) / 30 for x in zip(*all_test_results)]  # 对每个测试结果计算平均值

    # 创建Excel工作簿并写入结果
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for i, (combination, avg_result) in enumerate(zip(new_arr, avg_test_results)):
        worksheet.cell(row=i + 1, column=1, value=combination)
        worksheet.cell(row=i + 1, column=2, value=avg_result)

    workbook.save('../output/average_output_regtree_popt_30_log_perpopt_newData_loc.xlsx')